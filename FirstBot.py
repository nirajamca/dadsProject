from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.common.by import By
import time
import openpyxl
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import PatternFill, Font
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

correctMonth = False
correctYear = False

# Ask user for month
while correctMonth is not True:
    inputMonth = input("Month: ")
    # Select month user typed
    match inputMonth:
        case "JAN":
            correctMonth = True
        case "FEB":
            correctMonth = True
        case "MAR":
            correctMonth = True
        case "APR":
            correctMonth = True
        case "MAY":
            correctMonth = True
        case "JUN":
            correctMonth = True
        case "JUL":
            correctMonth = True
        case "AUG":
            correctMonth = True
        case "SEP":
            correctMonth = True
        case "OCT":
            correctMonth = True
        case "NOV":
            correctMonth = True
        case "DEC":
            correctMonth = True
        case _:
            print("Invalid Month. Please try again with format MON (first 3 letters of month)")
            continue

# Ask user for year
while correctYear is not True:
    inputYear = input("Year: ")
    if 2017 <= int(inputYear) <= 2023:
        correctYear = True
    else:
        print("Invalid Year. Please try again with format YYYY")
        continue

# Set driver to use Chrome
chromedriver_autoinstaller.install()
myDriver = webdriver.Chrome()
myDriver.maximize_window()

# Launch URL to open the desired website
myDriver.get("https://dwcdataportal.fldfs.com/ProofOfCoverage.aspx")

# First drop down menu, select "Expiration Date"
expirationDate = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddlSearchOptions"]/option[4]')
expirationDate.click()

# Select year user typed
match inputYear:
    case "2023":
        selectYear = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddYear"]/option[1]')
    case "2022":
        selectYear = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddYear"]/option[2]')
    case "2021":
        selectYear = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddYear"]/option[3]')
    case "2020":
        selectYear = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddYear"]/option[4]')
    case "2019":
        selectYear = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddYear"]/option[5]')
    case "2018":
        selectYear = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddYear"]/option[6]')
    case "2017":
        selectYear = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddYear"]/option[7]')

# Click respective year
selectYear.click()

# Select month user typed
match inputMonth:
    case "JAN":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[1]')
        mon = "01"
    case "FEB":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[2]')
        mon = "02"
    case "MAR":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[3]')
        mon = "03"
    case "APR":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[4]')
        mon = "04"
    case "MAY":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[5]')
        mon = "05"
    case "JUN":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[6]')
        mon = "06"
    case "JUL":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[7]')
        mon = "07"
    case "AUG":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[8]')
        mon = "08"
    case "SEP":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[9]')
        mon = "09"
    case "OCT":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[10]')
        mon = "10"
    case "NOV":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[11]')
        mon = "11"
    case "DEC":
        selectMonth = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddMonth"]/option[12]')
        mon = "12"

# Click respective month
selectMonth.click()

# Select FLORIDA WC JOINT ASSOC
selectInsurer = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_ddInsurer"]/option[393]')
selectInsurer.click()

# Click search button
clickSearch = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_btnSearch2"]')
clickSearch.click()

# Refresh page to grab the of the "Export to Excel" button
myDriver.refresh()

# Click on "Export to Excel" button
clickExportToExcel = myDriver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_Button2"]')
clickExportToExcel.click()

# Set timer for window to 15 seconds
time.sleep(15)

# Close window when done
myDriver.quit()

# Workbook object
wb_obj = openpyxl.load_workbook("/Users/richardthomas/Downloads/ProofOfCoverageReport.xlsx")

# Sheet object created
sheet_obj = wb_obj.active

# Assign titles to 6 rows
sheet_obj.cell(row=1, column=22).value = "Agent Name"
sheet_obj.cell(row=1, column=23).value = "Contact Name"
sheet_obj.cell(row=1, column=24).value = "Contact Title"
sheet_obj.cell(row=1, column=25).value = "Contact Email"
sheet_obj.cell(row=1, column=26).value = "Contact Phone"
sheet_obj.cell(row=1, column=27).value = "Comments"

# Change background of each cell
sheet_obj['V1'].font = Font(size=11, bold=True)
sheet_obj['V1'].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
sheet_obj['W1'].font = Font(size=11, bold=True)
sheet_obj['W1'].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
sheet_obj['X1'].font = Font(size=11, bold=True)
sheet_obj['X1'].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
sheet_obj['Y1'].font = Font(size=11, bold=True)
sheet_obj['Y1'].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
sheet_obj['Z1'].font = Font(size=11, bold=True)
sheet_obj['Z1'].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
sheet_obj['AA1'].font = Font(size=11, bold=True)
sheet_obj['AA1'].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')

# Save the file with desired filename
wb_obj.save(inputYear + '_' + mon + '_JUA_LEADS.xlsx')

# Email login and the email to send to
email_user = 'rthomasnegron@gmail.com'
email_password = 'onltkofdrfwxkrww'
email_send = 'richard@premierworkerscomp.com'

# Subject of the email
subject = 'JUA Leads ' + inputYear + ' ' + mon

# Assign each section
msg = MIMEMultipart()
msg['From'] = email_user
msg['To'] = email_send
msg['Subject'] = subject

# Body of email
body = 'Attached is the excel sheet'
msg.attach(MIMEText(body, 'plain'))

# Filename to attach
filename = inputYear + '_' + mon + '_JUA_LEADS.xlsx'
attachment = open(filename, 'rb')

part = MIMEBase('application', 'octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', "attachment; filename= " + filename)

# Make smtp port and login
msg.attach(part)
text = msg.as_string()
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(email_user, email_password)

# Send and quit server
server.sendmail(email_user, email_send, text)
server.quit()
